from __future__ import annotations

import argparse
import csv
import sys
from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
from typing import Optional, Tuple

from openpyxl import load_workbook


SHEET_NAME = "H010"

# Regras fixas (1-indexed no Excel)
COL_QTD = 10  # J
COL_VL_UNIT = 11  # K
COL_VL_ITEM = 12  # L
COL_VL_ITEM_IR = 17  # Q


def _to_decimal(value) -> Optional[Decimal]:
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        # float -> str para evitar 0.30000000000000004
        return Decimal(str(value))
    if isinstance(value, Decimal):
        return value
    if isinstance(value, str):
        v = value.strip()
        if not v:
            return None
        # Aceita número com vírgula
        v = v.replace(".", "").replace(",", ".") if ("," in v and v.count(",") == 1) else v
        try:
            return Decimal(v)
        except InvalidOperation:
            return None
    return None


def _quantize(d: Decimal, places: int) -> Decimal:
    q = Decimal("1").scaleb(-places)  # 10^-places
    return d.quantize(q, rounding=ROUND_HALF_UP)


@dataclass(frozen=True)
class RowChange:
    row: int
    qtd: Decimal
    old_unit: Decimal
    new_unit: Decimal
    new_total: Decimal


def _parse_percent(p: str) -> Decimal:
    d = _to_decimal(p)
    if d is None:
        raise argparse.ArgumentTypeError("Percentual inválido. Ex.: 10, -3.5, 0.8")
    return d


def _make_output_path(input_path: Path) -> Path:
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    return input_path.with_name(f"{input_path.stem}_AJUSTADO_{ts}{input_path.suffix}")


def ajustar_arquivo(
    input_path: Path,
    output_path: Path,
    percentual: Decimal,
    unit_places: int,
    total_places: int,
) -> Tuple[int, int, int, Path, Path]:
    if not input_path.exists():
        raise FileNotFoundError(f"Arquivo não encontrado: {input_path}")
    if input_path.suffix.lower() not in {".xlsx", ".xlsm"}:
        raise ValueError("Use um arquivo .xlsx ou .xlsm (planilha do Excel).")

    wb = load_workbook(filename=str(input_path), data_only=False, keep_vba=(input_path.suffix.lower() == ".xlsm"))

    if SHEET_NAME not in wb.sheetnames:
        raise KeyError(
            f"A aba obrigatória '{SHEET_NAME}' não existe. Abas encontradas: {', '.join(wb.sheetnames)}"
        )

    ws = wb[SHEET_NAME]

    fator = Decimal("1") + (percentual / Decimal("100"))

    changes: list[RowChange] = []
    skipped_empty = 0
    skipped_non_numeric = 0

    # Começa da linha 2 para preservar cabeçalho típico; se não houver, ainda funciona.
    for r in range(2, ws.max_row + 1):
        c_qtd = ws.cell(row=r, column=COL_QTD)
        c_unit = ws.cell(row=r, column=COL_VL_UNIT)

        # Ignora linhas totalmente vazias nessas colunas
        if c_qtd.value is None and c_unit.value is None:
            skipped_empty += 1
            continue

        # Não tenta mexer em fórmulas (openpyxl não calcula)
        if isinstance(c_qtd.value, str) and c_qtd.value.startswith("="):
            skipped_non_numeric += 1
            continue
        if isinstance(c_unit.value, str) and c_unit.value.startswith("="):
            skipped_non_numeric += 1
            continue

        qtd = _to_decimal(c_qtd.value)
        old_unit = _to_decimal(c_unit.value)

        if qtd is None or old_unit is None:
            skipped_non_numeric += 1
            continue

        new_unit = _quantize(old_unit * fator, unit_places)
        new_total = _quantize(qtd * new_unit, total_places)

        # Aplica regras:
        # 1) Ajusta apenas VL_UNIT (K)
        c_unit.value = float(new_unit)
        # 2) Recalcula VL_ITEM (L) = QTD (J) * novo VL_UNIT
        ws.cell(row=r, column=COL_VL_ITEM).value = float(new_total)
        # 3) VL_ITEM_IR (Q) = VL_ITEM
        ws.cell(row=r, column=COL_VL_ITEM_IR).value = float(new_total)

        changes.append(RowChange(row=r, qtd=qtd, old_unit=old_unit, new_unit=new_unit, new_total=new_total))

    # Log CSV ao lado do output
    log_path = output_path.with_suffix(".log.csv")
    with log_path.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f, delimiter=";")
        w.writerow(["row", "qtd(J)", "vl_unit_antigo(K)", "vl_unit_novo(K)", "vl_item_novo(L)", "vl_item_ir_novo(Q)"])
        for ch in changes:
            w.writerow(
                [
                    ch.row,
                    str(ch.qtd),
                    str(ch.old_unit),
                    str(ch.new_unit),
                    str(ch.new_total),
                    str(ch.new_total),
                ]
            )

    wb.save(str(output_path))
    return len(changes), skipped_empty, skipped_non_numeric, output_path, log_path


def main(argv: list[str]) -> int:
    p = argparse.ArgumentParser(
        description=(
            "Ajusta automaticamente a aba H010 aplicando um percentual único em VL_UNIT (K), "
            "recalculando VL_ITEM (L) e VL_ITEM_IR (Q)."
        )
    )
    p.add_argument("-i", "--input", required=True, help="Caminho do Excel de entrada (.xlsx/.xlsm).")
    p.add_argument(
        "-o",
        "--output",
        required=False,
        help="Caminho do Excel de saída. Se omitido, cria um arquivo *_AJUSTADO_YYYYMMDD_HHMMSS.xlsx ao lado.",
    )
    p.add_argument(
        "-p",
        "--percentual",
        required=True,
        type=_parse_percent,
        help="Percentual a aplicar em VL_UNIT. Ex.: 10 para +10%%, -5 para -5%%.",
    )
    p.add_argument(
        "--unit-places",
        type=int,
        default=6,
        help="Casas decimais para o VL_UNIT (K). Padrão: 6.",
    )
    p.add_argument(
        "--total-places",
        type=int,
        default=2,
        help="Casas decimais para VL_ITEM (L) e VL_ITEM_IR (Q). Padrão: 2.",
    )

    args = p.parse_args(argv)

    input_path = Path(args.input).expanduser().resolve()
    output_path = Path(args.output).expanduser().resolve() if args.output else _make_output_path(input_path)

    try:
        changed, skipped_empty, skipped_non_numeric, out_xlsx, out_log = ajustar_arquivo(
            input_path=input_path,
            output_path=output_path,
            percentual=args.percentual,
            unit_places=args.unit_places,
            total_places=args.total_places,
        )
    except Exception as e:
        print(f"ERRO: {e}", file=sys.stderr)
        return 2

    print("OK")
    print(f"- Arquivo de saída: {out_xlsx}")
    print(f"- Log (CSV): {out_log}")
    print(f"- Linhas ajustadas: {changed}")
    print(f"- Linhas ignoradas (vazias em J/K): {skipped_empty}")
    print(f"- Linhas ignoradas (não numéricas/fórmulas): {skipped_non_numeric}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))

