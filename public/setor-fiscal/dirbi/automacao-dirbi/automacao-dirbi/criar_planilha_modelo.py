"""
Script para criar a planilha modelo com códigos e naturezas de receita
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# Dicionário com códigos e naturezas da receita
CODIGOS_NATUREZAS = {
    "101": "ADUBOS E FERTILIZANTES",
    "102": "DEFENSIVOS AGRICOLAS",
    "103": "SEMENTES E MUDAS",
    "104": "CORRETIVO DE SOLO",
    "105": "FEIJOES, ARROZ, FARINHAS E SEMOLAS",
    "106": "INOCULANTES AGRICOLAS",
    "107": "VACINAS VETERINARIAS",
    "108": "FARINHAS A BASE DE MILHO",
    "110": "LEITE FLUIDO OU PASTEURIZADO, LEITE EM PO",
    "111": "QUEIJOS",
    "112": "SORO DE LEITE",
    "113": "FARINHA DE TRIGO",
    "114": "TRIGO",
    "115": "PRE MISTURA PARA PÃO",
    "119": "MASSAS ALIMENTICIAS",
    "121": "CARNES",
    "122": "PEIXES",
    "123": "CAFÉ",
    "124": "AÇUCAR",
    "125": "OLEOS VEGETAIS",
    "126": "MANTEIGA",
    "127": "MARGARINA",
    "128": "SABAO DE TOUCADOR",
    "129": "PRODUTOS DE HIGIENE BUCAL",
    "130": "PAPEL HIGIENICO",
    "116": "PRODUTOS HORTÍCOLAS E FRUTAS",
    "117": "OVOS",
    "118": "VENDA DE SÊMENS E EMBRIÕES",
    "301": "CADEIRAS DE RODAS E OUTROS VEÍCULOS",
    "302": "ARTIGOS E APARELHOS ORTOPÉDICOS OU PARA FRATURAS",
    "303": "ARTIGOS E APARELHOS DE PRÓTESES",
    "304": "ALMOFADAS ANTI ESCARAS",
    "306": "PRODUTOS QUIMICOS",
    "308": "PRODUTOS DESTINADOS AO USO EM HOSPITAIS, CLÍNICAS E CONSULTÓRIOS MÉDICOS E ODONTOLÓGICOS, CAMPANHAS DE SAÚDE REALIZADAS PELO PODER PÚBLICO, LABORATÓRIO DE ANATOMIA PATOLÓGICA, CITOLÓGICA OU DE ANÁLISES CLÍNICAS",
    "309": "PRODUTOS CLASSIFICADOS NOS CÓDIGOS 8443.32.22, 8469.00.39 EX 01, 8714.20.00, 9021.40.00, 9021.90.82 E 9021.90.92",
    "310": "CALCULADORAS EQUIPADAS COM SINTETIZADOR DE VOZ",
    "311": "TECLADOS COM COLMEIA",
    "312": "INDICADORES OU APONTADORES - MOUSES - COM ENTRADA PARA ACIONADOR",
    "313": "LINHAS BRAILE",
    "314": "DIGITALIZADORES DE IMAGENS - SCANNERS - EQUIPADOS COM SINTETIZADOR DE VOZ",
    "315": "DUPLICADORES BRAILE",
    "316": "ACIONADORES DE PRESSÃO",
    "317": "LUPAS ELETRÔNICAS DO TIPO UTILIZADO POR PESSOAS COM DEFICIÊNCIA VISUAL",
    "318": "IMPLANTES COCLEARES",
    "319": "PRÓTESES OCULARES",
    "320": "PROGRAMAS - SOFTWARES - DE LEITORES DE TELA QUE CONVERTEM TEXTO EM VOZ SINTETIZADA PARA AUXÍLIO DE PESSOAS COM DEFICIÊNCIA VISUAL",
    "321": "APARELHOS CONTENDO PROGRAMAS - SOFTWARES - DE LEITORES DE TELA QUE CONVERTEM TEXTO EM CARACTERES BRAILE, PARA UTILIZAÇÃO DE SURDOS-CEGOS",
    "322": "NEUROESTIMULADORES PARA TREMOR ESSENCIAL/PARKINSON",
    "903": "LIVROS",
    "914": "RECEITA DECORRENTE DA VENDA DE ÁGUAS MINERAIS NATURAIS COMERCIALIZADAS EM RECIPIENTES COM CAPACIDADE NOMINAL INFERIOR A 10 (DEZ) LITROS OU IGUAL OU SUPERIOR A 10 (DEZ) LITROS"
}

def criar_planilha_modelo():
    """Cria a planilha modelo com códigos e naturezas"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Receitas"
    
    # Estilo para cabeçalho
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Cabeçalhos
    headers = [
        "CÓDIGO DA RECEITA",
        "NATUREZA DA RECEITA",
        "VALOR DA RECEITA",
        "VALOR DA COFINS",
        "VALOR DO PIS"
    ]
    
    # Preenche cabeçalhos
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border
    
    # Ajusta largura das colunas
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18
    
    # Preenche dados
    linha = 2
    for codigo, natureza in sorted(CODIGOS_NATUREZAS.items(), key=lambda x: int(x[0])):
        # Código da receita
        cell_codigo = ws.cell(row=linha, column=1)
        cell_codigo.value = codigo
        cell_codigo.alignment = Alignment(horizontal="center")
        cell_codigo.border = border
        
        # Natureza da receita
        cell_natureza = ws.cell(row=linha, column=2)
        cell_natureza.value = natureza
        cell_natureza.alignment = Alignment(horizontal="left")
        cell_natureza.border = border
        
        # Valor da receita (vazio - será preenchido pela automação)
        cell_valor = ws.cell(row=linha, column=3)
        cell_valor.value = None
        cell_valor.number_format = '#,##0.00'
        cell_valor.alignment = Alignment(horizontal="right")
        cell_valor.border = border
        
        # Valor da COFINS (vazio - será calculado)
        cell_cofins = ws.cell(row=linha, column=4)
        cell_cofins.value = None
        cell_cofins.number_format = '#,##0.00'
        cell_cofins.alignment = Alignment(horizontal="right")
        cell_cofins.border = border
        
        # Valor do PIS (vazio - será calculado)
        cell_pis = ws.cell(row=linha, column=5)
        cell_pis.value = None
        cell_pis.number_format = '#,##0.00'
        cell_pis.alignment = Alignment(horizontal="right")
        cell_pis.border = border
        
        linha += 1
    
    # Salva a planilha
    nome_arquivo = "Planilha_Modelo_Receitas.xlsx"
    wb.save(nome_arquivo)
    print(f"Planilha modelo criada: {nome_arquivo}")
    return nome_arquivo

if __name__ == "__main__":
    try:
        nome_arquivo = criar_planilha_modelo()
        import os
        if os.path.exists(nome_arquivo):
            print(f"\n✓ Planilha criada com sucesso: {nome_arquivo}")
            print(f"  Tamanho: {os.path.getsize(nome_arquivo)} bytes")
        else:
            print(f"\n✗ Erro: Arquivo não foi criado")
    except Exception as e:
        print(f"\n✗ Erro ao criar planilha: {str(e)}")
        import traceback
        traceback.print_exc()

