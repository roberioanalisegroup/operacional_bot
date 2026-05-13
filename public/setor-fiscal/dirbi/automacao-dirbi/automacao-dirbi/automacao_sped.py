"""
Automação para importar SPED Contribuições e preencher planilha
Analisa registros M410 e extrai código da receita e valores
"""

import re
import os
import sys
from pathlib import Path
from tkinter import filedialog, messagebox, ttk
import tkinter as tk
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from collections import defaultdict


class AutomacaoSPED:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("Automação SPED Contribuições - M410")
        self.root.geometry("900x700")
        self.root.resizable(True, True)
        
        # Variáveis
        self.arquivo_sped = None
        self.dados_processados = None
        
        # Cria interface
        self.criar_interface()
        
    def criar_interface(self):
        """Cria a interface gráfica"""
        # Frame principal
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Configura grid
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        main_frame.columnconfigure(1, weight=1)
        main_frame.rowconfigure(2, weight=1)
        
        # Título
        titulo = ttk.Label(main_frame, text="Automação SPED Contribuições", 
                          font=("Arial", 16, "bold"))
        titulo.grid(row=0, column=0, columnspan=3, pady=(0, 20))
        
        # Seção 1: Seleção de arquivo SPED
        frame_sped = ttk.LabelFrame(main_frame, text="1. Selecionar Arquivo SPED", padding="10")
        frame_sped.grid(row=1, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=5)
        frame_sped.columnconfigure(1, weight=1)
        
        ttk.Button(frame_sped, text="Selecionar SPED", 
                  command=self.selecionar_arquivo_sped).grid(row=0, column=0, padx=5)
        
        self.label_arquivo = ttk.Label(frame_sped, text="Nenhum arquivo selecionado", 
                                       foreground="gray")
        self.label_arquivo.grid(row=0, column=1, sticky=(tk.W, tk.E), padx=5)
        
        ttk.Button(frame_sped, text="Processar", 
                  command=self.processar_arquivo).grid(row=0, column=2, padx=5)
        
        # Seção 2: Painel de informações
        frame_info = ttk.LabelFrame(main_frame, text="2. Dados Extraídos do SPED", padding="10")
        frame_info.grid(row=2, column=0, columnspan=3, sticky=(tk.W, tk.E, tk.N, tk.S), pady=5)
        frame_info.columnconfigure(0, weight=1)
        frame_info.rowconfigure(1, weight=1)
        
        # Frame para estatísticas
        frame_stats = ttk.Frame(frame_info)
        frame_stats.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=(0, 10))
        frame_stats.columnconfigure(1, weight=1)
        frame_stats.columnconfigure(3, weight=1)
        frame_stats.columnconfigure(5, weight=1)
        
        ttk.Label(frame_stats, text="Total de Registros:", font=("Arial", 10, "bold")).grid(row=0, column=0, padx=5)
        self.label_total_registros = ttk.Label(frame_stats, text="0", font=("Arial", 10))
        self.label_total_registros.grid(row=0, column=1, padx=5, sticky=tk.W)
        
        ttk.Label(frame_stats, text="Valor Total:", font=("Arial", 10, "bold")).grid(row=0, column=2, padx=5)
        self.label_valor_total = ttk.Label(frame_stats, text="R$ 0,00", font=("Arial", 10))
        self.label_valor_total.grid(row=0, column=3, padx=5, sticky=tk.W)
        
        ttk.Label(frame_stats, text="Códigos Encontrados:", font=("Arial", 10, "bold")).grid(row=0, column=4, padx=5)
        self.label_codigos = ttk.Label(frame_stats, text="0", font=("Arial", 10))
        self.label_codigos.grid(row=0, column=5, padx=5, sticky=tk.W)
        
        # Treeview para exibir dados
        columns = ("Código", "Valor Receita", "COFINS (7,6%)", "PIS (1,65%)")
        self.tree = ttk.Treeview(frame_info, columns=columns, show="headings", height=15)
        
        # Configura colunas
        self.tree.heading("Código", text="Código da Receita")
        self.tree.heading("Valor Receita", text="Valor da Receita")
        self.tree.heading("COFINS (7,6%)", text="COFINS (7,6%)")
        self.tree.heading("PIS (1,65%)", text="PIS (1,65%)")
        
        self.tree.column("Código", width=120, anchor=tk.CENTER)
        self.tree.column("Valor Receita", width=150, anchor=tk.E)
        self.tree.column("COFINS (7,6%)", width=150, anchor=tk.E)
        self.tree.column("PIS (1,65%)", width=150, anchor=tk.E)
        
        # Scrollbar
        scrollbar = ttk.Scrollbar(frame_info, orient=tk.VERTICAL, command=self.tree.yview)
        self.tree.configure(yscrollcommand=scrollbar.set)
        
        self.tree.grid(row=1, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        scrollbar.grid(row=1, column=1, sticky=(tk.N, tk.S))
        
        # Seção 3: Ações
        frame_acoes = ttk.LabelFrame(main_frame, text="3. Ações", padding="10")
        frame_acoes.grid(row=3, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=5)
        
        ttk.Button(frame_acoes, text="Criar Planilha Modelo", 
                  command=self.criar_planilha_modelo_interface).grid(row=0, column=0, padx=5)
        
        ttk.Button(frame_acoes, text="Exportar Planilha Preenchida", 
                  command=self.exportar_planilha, state=tk.DISABLED).grid(row=0, column=1, padx=5)
        self.btn_exportar = frame_acoes.winfo_children()[1]  # Referência ao botão exportar
        
        ttk.Button(frame_acoes, text="Sair", 
                  command=self.root.quit).grid(row=0, column=2, padx=5)
        
    def selecionar_arquivo_sped(self):
        """Abre diálogo para selecionar arquivo SPED"""
        arquivo = filedialog.askopenfilename(
            title="Selecione o arquivo SPED Contribuições",
            filetypes=[("Arquivos SPED", "*.txt"), ("Todos os arquivos", "*.*")]
        )
        if arquivo:
            self.arquivo_sped = arquivo
            nome_arquivo = os.path.basename(arquivo)
            self.label_arquivo.config(text=nome_arquivo, foreground="black")
    
    def parsear_registro_m400(self, linha):
        """
        Parseia uma linha do tipo M400
        Formato: |M400|código|valor|código||
        Retorna: código ou None
        """
        linha = linha.strip()
        
        if not linha.startswith('|M400|'):
            return None
        
        campos = linha.split('|')
        campos = [c for c in campos if c]
        
        if len(campos) >= 2 and campos[0] == 'M400':
            return campos[1].strip()
        
        return None
    
    def parsear_registro_m410(self, linha):
        """
        Parseia uma linha do tipo M410
        Formato: |M410|código_receita|valor|código||
        Retorna: (código_receita, valor) ou None
        """
        linha = linha.strip()
        
        if not linha.startswith('|M410|'):
            return None
        
        campos = linha.split('|')
        campos = [c for c in campos if c]
        
        if len(campos) >= 4 and campos[0] == 'M410':
            codigo_receita = campos[1].strip()
            valor_str = campos[2].strip()
            codigo = campos[3].strip() if len(campos) > 3 else ''
            
            try:
                valor = float(valor_str.replace(',', '.'))
                return (codigo_receita, valor, codigo)
            except ValueError:
                return None
        
        return None
    
    def processar_arquivo(self):
        """Processa o arquivo SPED selecionado - apenas M410 abaixo de M400|06|"""
        if not self.arquivo_sped:
            messagebox.showwarning("Aviso", "Por favor, selecione um arquivo SPED primeiro.")
            return
        
        try:
            dados = defaultdict(float)
            dentro_do_m400_06 = False
            encontrou_m400_06 = False
            
            with open(self.arquivo_sped, 'r', encoding='utf-8') as f:
                for linha in f:
                    # Verifica se é M400
                    codigo_m400 = self.parsear_registro_m400(linha)
                    if codigo_m400 is not None:
                        # Se encontrou M400|06|, ativa o processamento
                        if codigo_m400 == '06':
                            dentro_do_m400_06 = True
                            encontrou_m400_06 = True
                        else:
                            # Se encontrou outro M400, para o processamento
                            dentro_do_m400_06 = False
                        continue
                    
                    # Se está dentro do bloco M400|06|, processa M410
                    if dentro_do_m400_06:
                        resultado = self.parsear_registro_m410(linha)
                        if resultado:
                            codigo_receita, valor, codigo = resultado
                            dados[codigo_receita] += valor
            
            if not encontrou_m400_06:
                messagebox.showwarning("Aviso", "Nenhum registro |M400|06| encontrado no arquivo.")
                return
            
            if len(dados) == 0:
                messagebox.showwarning("Aviso", "Nenhum registro M410 encontrado abaixo de |M400|06|.")
                return
            
            self.dados_processados = dict(dados)
            self.atualizar_painel()
            self.btn_exportar.config(state=tk.NORMAL)
            messagebox.showinfo("Sucesso", f"Arquivo processado com sucesso!\n{len(dados)} código(s) encontrado(s) abaixo de |M400|06|.")
            
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao processar arquivo SPED:\n{str(e)}")
    
    def atualizar_painel(self):
        """Atualiza o painel com os dados processados"""
        if not self.dados_processados:
            return
        
        # Limpa treeview
        for item in self.tree.get_children():
            self.tree.delete(item)
        
        # Calcula totais
        total_registros = len(self.dados_processados)
        total_valor = sum(self.dados_processados.values())
        total_cofins = total_valor * 0.076
        total_pis = total_valor * 0.0165
        
        # Função auxiliar para formatar valores
        def formatar_moeda(valor):
            """Formata valor como moeda brasileira"""
            return f"R$ {valor:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
        
        # Atualiza labels
        self.label_total_registros.config(text=str(total_registros))
        self.label_valor_total.config(text=formatar_moeda(total_valor))
        self.label_codigos.config(text=str(total_registros))
        
        # Preenche treeview
        TAXA_PIS = 0.0165
        TAXA_COFINS = 0.076
        
        for codigo, valor in sorted(self.dados_processados.items()):
            valor_cofins = valor * TAXA_COFINS
            valor_pis = valor * TAXA_PIS
            
            self.tree.insert("", tk.END, values=(
                codigo,
                formatar_moeda(valor),
                formatar_moeda(valor_cofins),
                formatar_moeda(valor_pis)
            ))
        
        # Adiciona linha de total
        self.tree.insert("", tk.END, values=(
            "TOTAL",
            formatar_moeda(total_valor),
            formatar_moeda(total_cofins),
            formatar_moeda(total_pis)
        ), tags=("total",))
        self.tree.tag_configure("total", background="#E0E0E0", font=("Arial", 9, "bold"))
    
    def criar_planilha_modelo_interface(self):
        """Cria a planilha modelo via interface"""
        arquivo = filedialog.asksaveasfilename(
            title="Salvar Planilha Modelo",
            defaultextension=".xlsx",
            filetypes=[("Arquivos Excel", "*.xlsx"), ("Todos os arquivos", "*.*")]
        )
        
        if arquivo:
            try:
                nome_arquivo = self.criar_planilha_modelo(arquivo)
                messagebox.showinfo("Sucesso", f"Planilha modelo criada com sucesso!\n\n{arquivo}")
            except Exception as e:
                messagebox.showerror("Erro", f"Erro ao criar planilha modelo:\n{str(e)}")
    
    def criar_planilha_modelo(self, caminho_arquivo=None):
        """Cria a planilha modelo com códigos e naturezas"""
        CODIGOS_NATUREZAS = {
            "101": "ADUBOS E FERTILIZANTES",
            "102": "DEFENSIVOS AGRICOLAS",
            "103": "SEMENTES E MUDAS",
            "104": "CORRETIVO DE SOLO",
            "105": "FEIJOES, ARROZ, FARINHAS E SEMOLAS",
            "106": "INOCULANTES AGRICOLAS",
            "107": "VACINAS VETERINARIAS",
            "108": "FARINHAS A BASE DE MILHO",
            "110": "LEITE FLUIDO OU PASTEURIZADO, LEITE EM PO",
            "111": "QUEIJOS",
            "112": "SORO DE LEITE",
            "113": "FARINHA DE TRIGO",
            "114": "TRIGO",
            "115": "PRE MISTURA PARA PÃO",
            "119": "MASSAS ALIMENTICIAS",
            "121": "CARNES",
            "122": "PEIXES",
            "123": "CAFÉ",
            "124": "AÇUCAR",
            "125": "OLEOS VEGETAIS",
            "126": "MANTEIGA",
            "127": "MARGARINA",
            "128": "SABAO DE TOUCADOR",
            "129": "PRODUTOS DE HIGIENE BUCAL",
            "130": "PAPEL HIGIENICO",
            "116": "PRODUTOS HORTÍCOLAS E FRUTAS",
            "117": "OVOS",
            "118": "VENDA DE SÊMENS E EMBRIÕES",
            "301": "CADEIRAS DE RODAS E OUTROS VEÍCULOS",
            "302": "ARTIGOS E APARELHOS ORTOPÉDICOS OU PARA FRATURAS",
            "303": "ARTIGOS E APARELHOS DE PRÓTESES",
            "304": "ALMOFADAS ANTI ESCARAS",
            "306": "PRODUTOS QUIMICOS",
            "308": "PRODUTOS DESTINADOS AO USO EM HOSPITAIS, CLÍNICAS E CONSULTÓRIOS MÉDICOS E ODONTOLÓGICOS, CAMPANHAS DE SAÚDE REALIZADAS PELO PODER PÚBLICO, LABORATÓRIO DE ANATOMIA PATOLÓGICA, CITOLÓGICA OU DE ANÁLISES CLÍNICAS",
            "309": "PRODUTOS CLASSIFICADOS NOS CÓDIGOS 8443.32.22, 8469.00.39 EX 01, 8714.20.00, 9021.40.00, 9021.90.82 E 9021.90.92",
            "310": "CALCULADORAS EQUIPADAS COM SINTETIZADOR DE VOZ",
            "311": "TECLADOS COM COLMEIA",
            "312": "INDICADORES OU APONTADORES - MOUSES - COM ENTRADA PARA ACIONADOR",
            "313": "LINHAS BRAILE",
            "314": "DIGITALIZADORES DE IMAGENS - SCANNERS - EQUIPADOS COM SINTETIZADOR DE VOZ",
            "315": "DUPLICADORES BRAILE",
            "316": "ACIONADORES DE PRESSÃO",
            "317": "LUPAS ELETRÔNICAS DO TIPO UTILIZADO POR PESSOAS COM DEFICIÊNCIA VISUAL",
            "318": "IMPLANTES COCLEARES",
            "319": "PRÓTESES OCULARES",
            "320": "PROGRAMAS - SOFTWARES - DE LEITORES DE TELA QUE CONVERTEM TEXTO EM VOZ SINTETIZADA PARA AUXÍLIO DE PESSOAS COM DEFICIÊNCIA VISUAL",
            "321": "APARELHOS CONTENDO PROGRAMAS - SOFTWARES - DE LEITORES DE TELA QUE CONVERTEM TEXTO EM CARACTERES BRAILE, PARA UTILIZAÇÃO DE SURDOS-CEGOS",
            "322": "NEUROESTIMULADORES PARA TREMOR ESSENCIAL/PARKINSON",
            "903": "LIVROS",
            "914": "RECEITA DECORRENTE DA VENDA DE ÁGUAS MINERAIS NATURAIS COMERCIALIZADAS EM RECIPIENTES COM CAPACIDADE NOMINAL INFERIOR A 10 (DEZ) LITROS OU IGUAL OU SUPERIOR A 10 (DEZ) LITROS"
        }
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Receitas"
        
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        headers = [
            "CÓDIGO DA RECEITA",
            "NATUREZA DA RECEITA",
            "VALOR DA RECEITA",
            "VALOR DA COFINS",
            "VALOR DO PIS"
        ]
        
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border
        
        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 18
        
        linha = 2
        for codigo, natureza in sorted(CODIGOS_NATUREZAS.items(), key=lambda x: int(x[0])):
            ws.cell(row=linha, column=1).value = codigo
            ws.cell(row=linha, column=1).alignment = Alignment(horizontal="center")
            ws.cell(row=linha, column=1).border = border
            
            ws.cell(row=linha, column=2).value = natureza
            ws.cell(row=linha, column=2).alignment = Alignment(horizontal="left")
            ws.cell(row=linha, column=2).border = border
            
            for col in [3, 4, 5]:
                cell = ws.cell(row=linha, column=col)
                cell.value = None
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal="right")
                cell.border = border
            
            linha += 1
        
        if caminho_arquivo is None:
            caminho_arquivo = "Planilha_Modelo_Receitas.xlsx"
        
        wb.save(caminho_arquivo)
        return caminho_arquivo
    
    def exportar_planilha(self):
        """Exporta a planilha preenchida"""
        if not self.dados_processados:
            messagebox.showwarning("Aviso", "Por favor, processe um arquivo SPED primeiro.")
            return
        
        # Primeiro, seleciona a planilha modelo
        arquivo_planilha = filedialog.askopenfilename(
            title="Selecione a planilha modelo",
            filetypes=[("Arquivos Excel", "*.xlsx *.xls"), ("Todos os arquivos", "*.*")]
        )
        
        if not arquivo_planilha:
            return
        
        # Depois, escolhe onde salvar
        caminho_saida = filedialog.asksaveasfilename(
            title="Salvar Planilha Preenchida",
            defaultextension=".xlsx",
            filetypes=[("Arquivos Excel", "*.xlsx"), ("Todos os arquivos", "*.*")]
        )
        
        if not caminho_saida:
            return
        
        try:
            self.preencher_planilha(arquivo_planilha, caminho_saida, self.dados_processados)
            messagebox.showinfo("Sucesso", f"Planilha exportada com sucesso!\n\n{caminho_saida}")
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao exportar planilha:\n{str(e)}")
    
    def preencher_planilha(self, caminho_planilha, caminho_saida, dados):
        """
        Preenche a planilha com os dados extraídos
        dados: dict {código_receita: valor_total}
        """
        wb = load_workbook(caminho_planilha)
        ws = wb.active
        
        col_codigo = 1
        col_natureza = 2
        col_valor_receita = 3
        col_cofins = 4
        col_pis = 5
        
        linha_inicio = 2
        
        # Procura cabeçalhos
        for linha in range(1, min(3, ws.max_row + 1)):
            for col in range(1, min(6, ws.max_column + 1)):
                celula = ws.cell(row=linha, column=col)
                valor_celula = str(celula.value).upper() if celula.value else ""
                
                if "CÓDIGO" in valor_celula and "RECEITA" in valor_celula:
                    col_codigo = col
                    linha_inicio = linha + 1
                elif "NATUREZA" in valor_celula and "RECEITA" in valor_celula:
                    col_natureza = col
                elif "VALOR" in valor_celula and "RECEITA" in valor_celula:
                    col_valor_receita = col
                elif "COFINS" in valor_celula:
                    col_cofins = col
                elif "PIS" in valor_celula:
                    col_pis = col
        
        TAXA_PIS = 0.0165
        TAXA_COFINS = 0.076
        
        registros_preenchidos = 0
        codigos_nao_encontrados = []
        
        for codigo_receita, valor_total in dados.items():
            encontrado = False
            codigo_receita_str = str(codigo_receita).strip()
            
            for linha in range(linha_inicio, ws.max_row + 1):
                codigo_existente = ws.cell(row=linha, column=col_codigo).value
                
                if codigo_existente is None:
                    continue
                
                codigo_existente_str = str(codigo_existente).strip()
                
                try:
                    codigo_receita_int = int(codigo_receita_str)
                    codigo_existente_int = int(codigo_existente_str)
                    if codigo_receita_int == codigo_existente_int:
                        encontrado = True
                    else:
                        continue
                except ValueError:
                    if codigo_existente_str == codigo_receita_str:
                        encontrado = True
                    else:
                        continue
                
                if encontrado:
                    ws.cell(row=linha, column=col_valor_receita).value = valor_total
                    ws.cell(row=linha, column=col_valor_receita).number_format = '#,##0.00'
                    
                    valor_cofins = valor_total * TAXA_COFINS
                    ws.cell(row=linha, column=col_cofins).value = valor_cofins
                    ws.cell(row=linha, column=col_cofins).number_format = '#,##0.00'
                    
                    valor_pis = valor_total * TAXA_PIS
                    ws.cell(row=linha, column=col_pis).value = valor_pis
                    ws.cell(row=linha, column=col_pis).number_format = '#,##0.00'
                    
                    registros_preenchidos += 1
                    break
            
            if not encontrado:
                codigos_nao_encontrados.append(codigo_receita)
        
        wb.save(caminho_saida)
        
        if codigos_nao_encontrados:
            mensagem = f"Planilha exportada!\n\n"
            mensagem += f"Registros preenchidos: {registros_preenchidos}\n"
            mensagem += f"Códigos não encontrados: {len(codigos_nao_encontrados)}\n"
            mensagem += f"({', '.join(codigos_nao_encontrados[:5])}"
            if len(codigos_nao_encontrados) > 5:
                mensagem += f" e mais {len(codigos_nao_encontrados) - 5}..."
            mensagem += ")"
            messagebox.showwarning("Aviso", mensagem)
    
    def executar(self):
        """Executa a aplicação"""
        self.root.mainloop()


def main():
    app = AutomacaoSPED()
    app.executar()


if __name__ == "__main__":
    main()
